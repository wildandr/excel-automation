import openpyxl

# Fungsi untuk membuka file support dan mengambil data dari cell tertentu
def get_value_from_file(file_path, sheet_name, cell_range):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    data = []
    for row in ws[cell_range]:
        data.append([cell.value for cell in row])
    return data

# Fungsi utama untuk memperbarui sheet Rekon
def update_rekon_sheet(template_file, output_file):
    # Buka file template
    wb = openpyxl.load_workbook(template_file, data_only=False)
    ws_rekon = wb["Rekon IDR"]  # Asumsikan nama sheet adalah 'Rekon'

    # 1. Update Cell Raw 35-39 dari file Update Assumption_2019 NEW.xlsx
    assumption_data = get_value_from_file("Automation 1.1/Liquidity Gap/Support Files_BSM Stress Test 30082024_Final Template/Update Assumption_2019 NEW.xlsx", "Sheet1", "A35:A39")
    for i, value in enumerate(assumption_data, start=35):
        ws_rekon[f'A{i}'] = value[0]

    # 2. Update Cell U41-X97 dari BSM Stress Test 29082024_Final Template.xlsx
    rekon_idr_bau_idr = get_value_from_file("Automation 1.1/Liquidity Gap/Support Files_BSM Stress Test 30082024_Final Template/BSM Stress Test 29082024_Final Template.xlsx", "Rekon IDR", "U41:X97")
    for i, row in enumerate(rekon_idr_bau_idr, start=41):
        for j, value in enumerate(row, start=21):  # U adalah kolom 21
            ws_rekon.cell(row=i, column=j, value=value)

    # 3. Update Cell Z41-AC97 dari BSM Stress Test 29082024_Final Template.xlsx
    rekon_fcy_bau_fcy = get_value_from_file("Automation 1.1/Liquidity Gap/Support Files_BSM Stress Test 30082024_Final Template/BSM Stress Test 29082024_Final Template.xlsx", "Rekon IDR", "Z41:AC97")
    for i, row in enumerate(rekon_fcy_bau_fcy, start=41):
        for j, value in enumerate(row, start=26):  # Z adalah kolom 26
            ws_rekon.cell(row=i, column=j, value=value)

    # 4. Update Cell AG43-AJ51 dari Daily Report_12092024.xlsx
    daily_report_data_1 = get_value_from_file("Automation 1.1/Liquidity Gap/Support Files_BSM Stress Test 30082024_Final Template/Daily Report_12092024.xlsx", "F01", "AG43:AJ51")
    for i, row in enumerate(daily_report_data_1, start=43):
        for j, value in enumerate(row, start=33):  # AG adalah kolom 33
            ws_rekon.cell(row=i, column=j, value=value)

    # 5. Update Cell AG53-AJ61 dari PTBN_LNBR.xlsm
    ptbn_lnbr_data = get_value_from_file("Automation 1.1/Liquidity Gap/Support Files_BSM Stress Test 30082024_Final Template/PTBN_LNBR.xlsm", "IDR", "AG53:AJ61")
    for i, row in enumerate(ptbn_lnbr_data, start=53):
        for j, value in enumerate(row, start=33):
            ws_rekon.cell(row=i, column=j, value=value)

    # 6. Update Cell AG62-AJ86 dari Daily Report_12092024.xlsx
    daily_report_data_2 = get_value_from_file("Automation 1.1/Liquidity Gap/Support Files_BSM Stress Test 30082024_Final Template/Daily Report_12092024.xlsx", "F01", "AG62:AJ86")
    for i, row in enumerate(daily_report_data_2, start=62):
        for j, value in enumerate(row, start=33):
            ws_rekon.cell(row=i, column=j, value=value)

    # Simpan hasil ke file baru
    wb.save(output_file)
    print(f"File berhasil disimpan sebagai {output_file}")

# Jalankan fungsi automasi
template_file = "Automation 1.2/Liquidity Gap/Result_BSM Stress Test 30082024_Final Template/BSM Stress Test 30082024_Final Template.xlsx"
output_file = "Automation 1.2/Liquidity Gap/Result_BSM Stress Test 31082024_Final Template/BSM Stress Test 31082024_Final Template.xlsx"

update_rekon_sheet(template_file, output_file)
